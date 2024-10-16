import csv
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from .decorators import role_required
from django.core.paginator import Paginator
from .models import Downtime, Serveraccess, TechnicalSupportLog, TaskLog, ServerIp, CustomerForm
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment


@login_required
def adminpanel(request):
    return render(request, 'adminpanel.html')

@login_required
@role_required(allowed_roles=['AllRole','LogUser'])
def downtime_list(request):
    downtimes = Downtime.objects.all().order_by('-datetime')
    #serveraccesses = Serveraccess.objects.all().order_by('-datetime')

    # Pagination for Downtime logs
    downtime_paginator = Paginator(downtimes, 20)
    downtime_page_number = request.GET.get('downtime_page')
    downtime_page_obj = downtime_paginator.get_page(downtime_page_number)

    # Pagination for Server Access logs
    #serveraccess_paginator = Paginator(serveraccesses, 5)
    #serveraccess_page_number = request.GET.get('serveraccess_page')
    #serveraccess_page_obj = serveraccess_paginator.get_page(serveraccess_page_number)

    context = {
        'downtime_page_obj': downtime_page_obj,
        #'serveraccess_page_obj': serveraccess_page_obj,
    }
    return render(request, 'downtime_list.html', context)

@login_required
@role_required(allowed_roles=['AllRole','LogUser'])
def export_downtimes_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="downtimes.csv"'

    writer = csv.writer(response)
    writer.writerow(['SN', 'Date and Time', 'System/Channel Name', 'Reason', 'Downtime Total Time', 'Impact'])

    downtimes = Downtime.objects.all().order_by('-datetime')
    for downtime in downtimes:
        writer.writerow([downtime.sn, downtime.datetime, downtime.system_channel_name, downtime.reason, downtime.downtime_total_time, downtime.impact])

    return response

@login_required
@role_required(allowed_roles=['AllRole','LogUser'])
def export_serveraccess_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="serveraccess.csv"'

    writer = csv.writer(response)
    writer.writerow(['SN', 'Date and Time', 'Server Name', 'Purpose', 'Access by', 'Provided by', 'Time'])

    serveraccesses = Serveraccess.objects.all().order_by('-datetime')
    for log in serveraccesses:
        writer.writerow([log.sn, log.datetime, log.server_name, log.purpose, log.access_by, log.provided_by, log.time])

    return response

@login_required
@role_required(allowed_roles=['AllRole','LogUser'])
def serverlog_list(request):
    serveraccesses = Serveraccess.objects.all().order_by('-datetime')

    # Pagination for Server Access logs
    serveraccess_paginator = Paginator(serveraccesses, 20)
    serveraccess_page_number = request.GET.get('serveraccess_page')
    serveraccess_page_obj = serveraccess_paginator.get_page(serveraccess_page_number)

    context = {
        'serveraccess_page_obj': serveraccess_page_obj,
    }
    return render(request, 'serverlog.html', context)

@login_required
@role_required(allowed_roles=['AllRole', 'LogUser'])
def technical_log_list(request):
    tech_logs = TechnicalSupportLog.objects.all().order_by('-date')
    paginator = Paginator(tech_logs, 20)  # Show 5 logs per page.
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {'page_obj': page_obj}
    return render(request, 'technicallog.html', context)

@login_required
@role_required(allowed_roles=['AllRole', 'LogUser'])
def export_technical_logs_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="technical_support_logs.csv"'

    writer = csv.writer(response)
    writer.writerow(['SN', 'Issue No', 'Date', 'Category', 'Requested by', 'Issue logged to', 'Issue Description', 'Reference code', 'Handling person', 'Status'])

    tech_logs = TechnicalSupportLog.objects.all()
    for log in tech_logs:
        writer.writerow([log.sn, log.issue_no, log.date, log.category, log.requested_by, log.issue_logged_to, log.issue_description, log.reference_code, log.handling_person, log.status])

    return response


@login_required
@role_required(allowed_roles=['AllRole'])
def tasklog_list(request):
    selected_category = request.GET.get('category', 'Overall')

    if selected_category == 'Overall':
        tasklogs = TaskLog.objects.exclude(category__in=['Board 1', 'Board 2']).order_by('id')
    else:
        tasklogs = TaskLog.objects.filter(category=selected_category).order_by('id')

    paginator = Paginator(tasklogs, 50)  # Show 5 logs per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'tasklog_page_obj': page_obj,
        'selected_category': selected_category,
    }
    return render(request, 'tasklog.html', context)

def redirect_to_overall(request):
    return redirect('/adminpanel/tasklog/list/?category=Overall')


@login_required
@role_required(allowed_roles=['AllRole'])
def serverip(request):
    serverips = ServerIp.objects.all().order_by('-id')
    # return JsonResponse({'serverips': list(serverips.values())})
    paginator = Paginator(serverips, 50)  # Show 5 logs per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'serverip_page_obj': page_obj,
    }
    return render(request, 'serverip.html', context)


@login_required
@role_required(allowed_roles=['AllRole'])
def export_serverdetails_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="serverdetails.csv"'

    writer = csv.writer(response)
    writer.writerow(['SN', 'Server Name', 'Server IP', 'URL', 'Storage', 'Internet Access', 'Request By', 'Server OS', 'Database IP','Create at', 'Remarks'])

    # serveraccesses = Serveraccess.objects.all().order_by('-datetime')
    serverdetails = ServerIp.objects.all()

    for log in serverdetails:
        writer.writerow([log.id, log.server_name, log.server_ip, log.url, log.storage, log.internet_access, log.request_by,log.server_os,log.database_ip,log.create_at,log.remarks])

    return response


#branch feedback form views --- bulk-form.html

def bulk_form(request):
    if request.method == "POST":
        customer_name = request.POST.get('customer_name')
        customer_id = request.POST.get('customer_id')
        account_number = request.POST.get('account_number')
        branch = request.POST.get('branch')
        charge_deducted = request.POST.get('charge_deducted')
        customer_profile = request.POST.get('customer_profile')
        remarks = request.POST.get('remarks')

        CustomerForm.objects.create(
            customer_name=customer_name,
            customer_id=customer_id,
            account_number=account_number,
            branch=branch,
            charge_deducted=charge_deducted,
            customer_profile=customer_profile,
            remarks=remarks
        )
        messages.success(request, "You have successfully submitted the data.")
        return redirect('bulk_form')

    return render(request, 'bulk-form.html')

@login_required
def bulk_form_admin(request):
    form_list = CustomerForm.objects.all().order_by('-id')
    paginator = Paginator(form_list, 20)  # Show 20 entries per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'bulk-form-admin.html', {'page_obj': page_obj})

@login_required
def export_customer_form_csv(request):
    # Create an in-memory workbook and worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Customer Forms'

    # Define the titles for the columns, including SN and date_of_entry
    columns = ['SN', 'Customer Name', 'Customer ID', 'Account Number', 'Branch', 'Charge Deducted',
               'Customer Profile', 'Remarks', 'Date of Entry']
    row_num = 1

    # Assign the titles for each cell in the header row
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Iterate through the data and write it out row by row
    forms = CustomerForm.objects.all()

    for form in forms:
        row_num += 1
        worksheet.cell(row=row_num, column=1, value=form.id)  # SN as ID
        worksheet.cell(row=row_num, column=2, value=form.customer_name)
        worksheet.cell(row=row_num, column=3, value=form.customer_id)
        worksheet.cell(row=row_num, column=4, value=form.account_number)  # Account number will retain leading zeros
        worksheet.cell(row=row_num, column=5, value=form.branch)
        worksheet.cell(row=row_num, column=6, value=form.charge_deducted)
        worksheet.cell(row=row_num, column=7, value=form.customer_profile)
        worksheet.cell(row=row_num, column=8, value=form.remarks)

        # Write date_of_entry without timezone
        date_of_entry = form.date_of_entry.replace(tzinfo=None)  # Remove timezone info
        worksheet.cell(row=row_num, column=9, value=date_of_entry)  # Date of Entry

    # Adjust column widths
    for col_num, column_title in enumerate(columns, 1):
        column_letter = get_column_letter(col_num)
        worksheet.column_dimensions[column_letter].width = len(column_title) + 5

    # Create a response object and specify content type
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=customer_forms.xlsx'

    # Save the workbook to the response
    workbook.save(response)

    return response